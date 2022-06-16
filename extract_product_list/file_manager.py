import pandas as pd 
import openpyxl as op
import glob
from .errors import ExcelFileNoFound, ExcelContentNoSupported

def getExcelFile(folderPath):
    # Return first excel file
    excelFile = glob.glob(f'{folderPath}/*.xlsx')
    if excelFile: 
        return excelFile[0]
    raise ExcelFileNoFound(f'No se encontro ningun archivo excel en {folderPath}')

def readExcelFile(excelFile):
    listAmazonProducts = pd.read_excel(f'{excelFile}', engine='openpyxl', header=0)
    acceptedColumnName = 'AMAZON PRODUCT'
    asinProducts = []
    if listAmazonProducts.columns.values and (acceptedColumnName in listAmazonProducts.columns.values):
        for i in listAmazonProducts.index:
            try:
                asinProducts.append(str(listAmazonProducts.loc[i,acceptedColumnName]))
            except:
                pass
        return asinProducts
    raise ExcelContentNoSupported('El contenido del archivo no es el adecuado.')

def writeExcelFile(excelFile, amazonProduct):
    sheet = excelFile.active
    # last row of sheet
    lastRow = sheet.max_row
    # Write amazon product information
    sheet.cell(row = lastRow+1, column = 1).value = amazonProduct.getName
    sheet.cell(row = lastRow+1, column = 2).value = amazonProduct.getDescription
    sheet.cell(row = lastRow+1, column = 3).value = amazonProduct.getPrice
    sheet.cell(row = lastRow+1, column = 4).value = amazonProduct.getAvailability
    sheet.cell(row = lastRow+1, column = 5).value = amazonProduct.getUrl

def saveExcelFile(excelFile, folderPath):
    excelFile.save(f'{folderPath}\\amazon_search.xlsx')

def createExcelFile():
    # Create a new excel file
    excelFile = op.Workbook()
    sheet = excelFile.active
    # Headers
    sheet.cell(row = 1, column = 1).value = "NOMBRE"
    sheet.cell(row = 1, column = 2).value = "DESCRIPCION"
    sheet.cell(row = 1, column = 3).value = "PRECIO"
    sheet.cell(row = 1, column = 4).value = "ESTADO"
    sheet.cell(row = 1, column = 4).value = "URL"
    # Cells widht
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 40
    sheet.column_dimensions['E'].width = 40
    return excelFile